import openpyxl, datetime
import sys, os

date = datetime.date.today()
if getattr(sys, 'frozen', False):
    base_path = 'logs\\'
else:
    base_path = os.path.dirname(__file__)
path = f'{base_path}\\{date}.xlsx'

index = {'main' : 3, 'additionally' : 3, 'reserve' : 3, '150' : 3}

def create():
    global date
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "data"
    sheet.append(["Основной цех", "", "", "", "", "Резерв (завтра)", "", "", "", "", "Дополнительный цех", "", "", "", "150 цех"])
    sheet.append(["Класс", "Ширина полос, мм", "Кол-во роликов", "Станок", "", "Класс", "Ширина полос, мм", "Кол-во роликов", "Станок", "", "Класс", "Ширина полос, мм", "Кол-во роликов", "", "Класс", "Ширина полос, мм", "Кол-во роликов"])

    workbook.save(path)

def adjust_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Добавляем немного отступа
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

def result(log, stage, time):
    global date, index
    name = {1: "Станок 1", 2: "Lt-245", 3: "ИРИС"}

    workbook = openpyxl.load_workbook(path)
    sheet = workbook["data"]
    if stage == 'main':
        sheet[f"A{index[stage]}"] = log[0]
        sheet[f"B{index[stage]}"] = log[1]
        sheet[f"C{index[stage]}"] = log[2]
        sheet[f"D{index[stage]}"] = name[log[3]]
        # if time > 630:
        #     sheet[f"A{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        #     sheet[f"B{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        #     sheet[f"C{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        #     sheet[f"D{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    elif stage == 'reserve':
        sheet[f"F{index[stage]}"] = log[0]
        sheet[f"G{index[stage]}"] = log[1]
        sheet[f"H{index[stage]}"] = log[2]
        sheet[f"I{index[stage]}"] = name[log[3]]
        # if time > 630:
        #     sheet[f"F{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        #     sheet[f"G{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        #     sheet[f"H{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        #     sheet[f"I{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    elif stage == 'additionally':
        sheet[f"K{index[stage]}"] = log[0]
        sheet[f"L{index[stage]}"] = log[1]
        sheet[f"M{index[stage]}"] = log[2]
        # if time > 630:
        #     sheet[f"K{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        #     sheet[f"L{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        #     sheet[f"M{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
    else: 
        sheet[f"O{index[stage]}"] = log[0]
        sheet[f"P{index[stage]}"] = log[1]
        sheet[f"Q{index[stage]}"] = log[2]
        # if time > 630:
        #     sheet[f"O{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        #     sheet[f"P{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        #     sheet[f"Q{index[stage]}"].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
    index[stage] += 1

    adjust_column_width(sheet)
    workbook.save(path)  
